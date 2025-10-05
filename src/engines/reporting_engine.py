"""
Reporting Engine
Generates reports from diagnostic results and device data
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.device import Device
from models.diagnostic_result import DiagnosticResult, Severity
from models.backup_record import BackupRecord


class ReportingEngine:
    """Generates various reports"""
    
    def __init__(self, config: dict):
        self.logger = logging.getLogger(__name__)
        self.config = config
        
        # Get report directory from config
        report_dir = config.get('reporting', {}).get('directory', 'reports')
        self.report_dir = Path(report_dir)
        self.report_dir.mkdir(exist_ok=True, parents=True)
        
        self.company_name = config.get('reporting', {}).get('company_name', 'Network Engineering Team')
    
    def generate_network_health_report(
        self,
        devices: List[Device],
        diagnostic_results: List[DiagnosticResult],
        format: str = 'excel'
    ) -> Optional[Path]:
        """
        Generate comprehensive network health report.
        
        Args:
            devices: List of devices
            diagnostic_results: List of diagnostic results
            format: Output format ('excel', 'csv', 'html')
            
        Returns:
            Path to generated report file
        """
        self.logger.info(f"Generating network health report in {format} format")
        
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"network_health_report_{timestamp}"
            
            if format == 'excel':
                return self._generate_excel_report(devices, diagnostic_results, filename)
            elif format == 'csv':
                return self._generate_csv_report(devices, diagnostic_results, filename)
            else:
                self.logger.error(f"Unsupported format: {format}")
                return None
        
        except Exception as e:
            self.logger.error(f"Error generating report: {e}", exc_info=True)
            return None
    
    def _generate_excel_report(
        self,
        devices: List[Device],
        diagnostic_results: List[DiagnosticResult],
        filename: str
    ) -> Path:
        """Generate Excel report with multiple sheets"""
        
        filepath = self.report_dir / f"{filename}.xlsx"
        
        # Create Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            
            # Sheet 1: Executive Summary
            self._create_summary_sheet(writer, devices, diagnostic_results)
            
            # Sheet 2: Device Inventory
            self._create_device_inventory_sheet(writer, devices)
            
            # Sheet 3: Health Status
            self._create_health_status_sheet(writer, diagnostic_results)
            
            # Sheet 4: Critical Issues
            self._create_issues_sheet(writer, diagnostic_results, severity=Severity.CRITICAL)
            
            # Sheet 5: Warnings
            self._create_issues_sheet(writer, diagnostic_results, severity=Severity.WARNING, sheet_name='Warnings')
        
        # Apply formatting
        self._format_excel_report(filepath)
        
        self.logger.info(f"Excel report generated: {filepath}")
        return filepath
    
    def _create_summary_sheet(
        self,
        writer: pd.ExcelWriter,
        devices: List[Device],
        diagnostic_results: List[DiagnosticResult]
    ) -> None:
        """Create executive summary sheet"""
        
        # Calculate statistics
        total_devices = len(devices)
        reachable_devices = sum(1 for d in devices if d.is_reachable())
        
        critical_issues = sum(
            len(r.get_critical_issues()) for r in diagnostic_results
        )
        warning_issues = sum(
            len(r.get_warning_issues()) for r in diagnostic_results
        )
        
        # Health percentage
        devices_with_issues = len([r for r in diagnostic_results if r.has_issues()])
        health_percentage = ((total_devices - devices_with_issues) / total_devices * 100) if total_devices > 0 else 0
        
        # Create summary data
        summary_data = {
            'Metric': [
                'Report Generated',
                'Company',
                '',
                'Total Devices',
                'Reachable Devices',
                'Unreachable Devices',
                '',
                'Critical Issues',
                'Warnings',
                '',
                'Network Health Score'
            ],
            'Value': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                self.company_name,
                '',
                total_devices,
                reachable_devices,
                total_devices - reachable_devices,
                '',
                critical_issues,
                warning_issues,
                '',
                f"{health_percentage:.1f}%"
            ]
        }
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_device_inventory_sheet(self, writer: pd.ExcelWriter, devices: List[Device]) -> None:
        """Create device inventory sheet"""
        
        inventory_data = []
        for device in devices:
            inventory_data.append({
                'IP Address': device.ip_address,
                'Hostname': device.hostname or 'N/A',
                'Vendor': device.vendor or 'Unknown',
                'Device Type': device.device_type.value if device.device_type else 'Unknown',
                'Model': device.model or 'N/A',
                'OS Version': device.os_version or 'N/A',
                'Status': device.status.value,
                'Last Seen': device.last_seen.strftime('%Y-%m-%d %H:%M:%S') if device.last_seen else 'N/A',
                'Uptime': device.uptime or 'N/A'
            })
        
        df = pd.DataFrame(inventory_data)
        df.to_excel(writer, sheet_name='Device Inventory', index=False)
    
    def _create_health_status_sheet(
        self,
        writer: pd.ExcelWriter,
        diagnostic_results: List[DiagnosticResult]
    ) -> None:
        """Create health status sheet"""
        
        health_data = []
        for result in diagnostic_results:
            health_data.append({
                'Device IP': result.device_ip,
                'Hostname': result.device_hostname or 'N/A',
                'Workflow': result.workflow_name,
                'Timestamp': result.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'Overall Severity': result.overall_severity.value,
                'Critical Issues': len(result.get_critical_issues()),
                'Warnings': len(result.get_warning_issues()),
                'Total Issues': len(result.issues),
                'Execution Time (s)': f"{result.execution_time:.2f}",
                'Success': 'Yes' if result.success else 'No'
            })
        
        df = pd.DataFrame(health_data)
        df.to_excel(writer, sheet_name='Health Status', index=False)
    
    def _create_issues_sheet(
        self,
        writer: pd.ExcelWriter,
        diagnostic_results: List[DiagnosticResult],
        severity: Severity,
        sheet_name: str = None
    ) -> None:
        """Create issues sheet for specific severity"""
        
        if sheet_name is None:
            sheet_name = f'{severity.value.title()} Issues'
        
        issues_data = []
        for result in diagnostic_results:
            issues = [i for i in result.issues if i.severity == severity]
            
            for issue in issues:
                issues_data.append({
                    'Device IP': result.device_ip,
                    'Hostname': result.device_hostname or 'N/A',
                    'Timestamp': result.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                    'Issue Type': issue.type,
                    'Severity': issue.severity.value,
                    'Description': issue.description,
                    'Recommendation': issue.recommendation or 'N/A'
                })
        
        df = pd.DataFrame(issues_data)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _format_excel_report(self, filepath: Path) -> None:
        """Apply formatting to Excel report"""
        
        from openpyxl import load_workbook
        
        wb = load_workbook(filepath)
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Apply borders to all cells
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
        
        wb.save(filepath)
    
    def _generate_csv_report(
        self,
        devices: List[Device],
        diagnostic_results: List[DiagnosticResult],
        filename: str
    ) -> Path:
        """Generate CSV report"""
        
        filepath = self.report_dir / f"{filename}.csv"
        
        # Combine all data into single CSV
        report_data = []
        
        for device in devices:
            # Find diagnostics for this device
            device_results = [r for r in diagnostic_results if r.device_ip == device.ip_address]
            
            for result in device_results:
                report_data.append({
                    'IP Address': device.ip_address,
                    'Hostname': device.hostname or 'N/A',
                    'Vendor': device.vendor or 'Unknown',
                    'Status': device.status.value,
                    'Workflow': result.workflow_name,
                    'Severity': result.overall_severity.value,
                    'Critical Issues': len(result.get_critical_issues()),
                    'Warnings': len(result.get_warning_issues()),
                    'Timestamp': result.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                })
        
        df = pd.DataFrame(report_data)
        df.to_csv(filepath, index=False)
        
        self.logger.info(f"CSV report generated: {filepath}")
        return filepath
    
    def generate_backup_report(self, backup_records: List[BackupRecord]) -> Optional[Path]:
        """Generate report for backup operations"""
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = self.report_dir / f"backup_report_{timestamp}.xlsx"
        
        # Prepare data
        backup_data = []
        for record in backup_records:
            backup_data.append({
                'Device IP': record.device_ip,
                'Hostname': record.device_hostname or 'N/A',
                'Config Type': record.config_type,
                'Timestamp': record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'Status': 'Success' if record.backup_successful else 'Failed',
                'File Size (bytes)': record.size_bytes if record.backup_successful else 0,
                'File Path': str(record.file_path) if record.backup_successful else 'N/A',
                'Error': record.error_message or 'None'
            })
        
        df = pd.DataFrame(backup_data)
        df.to_excel(filepath, index=False, sheet_name='Backup Results')
        
        self.logger.info(f"Backup report generated: {filepath}")
        return filepath
